from flask import Flask, render_template, request, send_file, jsonify
import re
import requests
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from unidecode import unidecode
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

app = Flask(__name__)

EXCEL_FILE = os.environ.get("EXCEL_FILE", "nrl.xlsx")
MAX_WORKERS = int(os.environ.get("MAX_WORKERS", 20))
REQUEST_TIMEOUT = 8

_cached_docs = None

# Pre-compile regex patterns
RE_DOC_ID = re.compile(r'/d/([a-zA-Z0-9_-]+)')
RE_STT = re.compile(r'^\d{1,5}$')
RE_STT_FLEXIBLE = re.compile(r'^(\d{1,5})[.\)\]\s]*$')  # Match: "32", "32.", "32)", "32]"
RE_NRL = re.compile(r'^(\d+\.?\d*)$')
# Pattern cho MSSV (thường 8-12 số)
RE_MSSV = re.compile(r'\b\d{8,12}\b')
# Pattern cho dòng bảng (chứa nhiều cột)
RE_TABLE_ROW = re.compile(r'[\t|]|(?:\s{2,})')


def get_doc_links():
    global _cached_docs
    if _cached_docs is not None:
        return _cached_docs
    
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] File {EXCEL_FILE} khong ton tai!")
        return []
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        doc_links = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink and cell.hyperlink.target:
                    link = cell.hyperlink.target
                    if "docs.google.com/document" in link:
                        cell_value = str(cell.value) if cell.value else ""
                        doc_links.append({"link": link, "name": cell_value})
        
        seen = set()
        unique_docs = []
        for doc in doc_links:
            if doc["link"] not in seen:
                seen.add(doc["link"])
                unique_docs.append(doc)
        
        _cached_docs = unique_docs
        print(f"[INFO] Loaded {len(unique_docs)} docs from Excel")
        return unique_docs
    except Exception as e:
        print(f"[ERROR] Load Excel failed: {e}")
        return []


def read_doc_text(url, session):
    """Đọc nội dung Google Docs với retry"""
    try:
        match = RE_DOC_ID.search(url)
        if not match:
            return None
        doc_id = match.group(1)
        export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=txt"
        
        # Retry 2 lần
        for attempt in range(2):
            try:
                r = session.get(export_url, timeout=REQUEST_TIMEOUT)
                if r.status_code == 200 and "accounts.google.com" not in r.url:
                    return r.text
            except:
                if attempt == 0:
                    continue
        return None
    except:
        return None


def normalize_text(text):
    text = unidecode(text.lower())
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def is_valid_stt(s):
    """Kiểm tra STT - hỗ trợ nhiều format: 32, 32., 32), etc."""
    s = s.strip()
    # Thử match chính xác trước
    if RE_STT.match(s):
        return True
    # Thử match flexible (có dấu chấm, ngoặc cuối)
    if RE_STT_FLEXIBLE.match(s):
        return True
    return False


def extract_stt_value(s):
    """Trích xuất giá trị số từ STT"""
    s = s.strip()
    match = RE_STT.match(s)
    if match:
        return int(s)
    match = RE_STT_FLEXIBLE.match(s)
    if match:
        return int(match.group(1))
    return None


def is_valid_nrl(s):
    match = RE_NRL.match(s)
    if match:
        val = float(match.group(1))
        if 0 <= val <= 10:
            return True, val
    return False, None


def parse_table_row(line):
    """Tách dòng bảng thành các cột"""
    # Thử tách theo tab trước
    if '\t' in line:
        return [c.strip() for c in line.split('\t') if c.strip()]
    # Thử tách theo pipe |
    if '|' in line:
        return [c.strip() for c in line.split('|') if c.strip()]
    # Thử tách theo nhiều space (2+)
    parts = re.split(r'\s{2,}', line)
    if len(parts) > 1:
        return [c.strip() for c in parts if c.strip()]
    # Fallback: tách theo space đơn
    return [c.strip() for c in line.split() if c.strip()]


def find_nrl_in_parts(parts, mssv):
    """Tìm NRL trong các phần của dòng, ưu tiên phần sau MSSV"""
    mssv_idx = -1
    # Tìm vị trí MSSV
    for i, part in enumerate(parts):
        if mssv in part:
            mssv_idx = i
            break
    
    # Ưu tiên tìm NRL SAU MSSV (vì thường format: STT | Tên | Lớp | MSSV | NRL)
    candidates = []
    for i, part in enumerate(parts):
        part_clean = part.replace(',', '.').strip()
        valid, val = is_valid_nrl(part_clean)
        if valid and part_clean != mssv and len(part_clean) <= 4:  # NRL thường ngắn
            # Tính khoảng cách từ MSSV
            distance = abs(i - mssv_idx) if mssv_idx >= 0 else i
            # Ưu tiên sau MSSV
            priority = 0 if i > mssv_idx else 1
            candidates.append((priority, distance, val))
    
    if candidates:
        # Sắp xếp theo priority, rồi distance
        candidates.sort(key=lambda x: (x[0], x[1]))
        return candidates[0][2]
    return None


def find_stt_in_parts(parts, mssv):
    """Tìm STT trong các phần của dòng, ưu tiên phần đầu"""
    for i, part in enumerate(parts):
        if is_valid_stt(part) and part != mssv:
            # STT thường ở đầu và là số nhỏ
            val = extract_stt_value(part)
            if val is not None and val <= 10000:  # Giới hạn STT hợp lý
                return val
    return None


def find_stt_in_line(line, mssv):
    """Tìm STT trong một dòng - thử nhiều cách"""
    # Cách 1: Parse như bảng
    parts = parse_table_row(line)
    stt = find_stt_in_parts(parts, mssv)
    if stt:
        return stt
    
    # Cách 2: Tìm số đầu tiên trong dòng (thường là STT)
    match = re.match(r'^\s*(\d{1,5})[.\)\]\s]', line)
    if match:
        val = int(match.group(1))
        if val <= 10000 and str(val) != mssv:
            return val
    
    # Cách 3: Tìm số đứng đầu dòng sau khi strip
    stripped = line.strip()
    first_word = stripped.split()[0] if stripped.split() else ""
    if is_valid_stt(first_word) and first_word != mssv:
        val = extract_stt_value(first_word)
        if val and val <= 10000:
            return val
    
    return None


def find_student_in_content(content, ten_sv, mssv):
    """
    Tìm sinh viên với thuật toán cải tiến:
    1. Kiểm tra MSSV chính xác (word boundary)
    2. Ưu tiên dòng có CẢ tên VÀ MSSV
    3. Xử lý nhiều format bảng
    4. Tìm trong vùng lân cận nếu không cùng dòng
    """
    content_normalized = normalize_text(content)
    ten_normalized = normalize_text(ten_sv)
    
    # Tách họ tên thành các từ để tìm chính xác hơn
    ten_parts = ten_normalized.split()
    ten_cuoi = ten_parts[-1] if ten_parts else ten_normalized  # Tên riêng (từ cuối)
    
    # Kiểm tra MSSV với word boundary (tránh match một phần)
    mssv_pattern = re.compile(r'\b' + re.escape(mssv) + r'\b')
    if not mssv_pattern.search(content):
        return False, None, None
    
    # Kiểm tra tên có trong content không
    if ten_normalized not in content_normalized and ten_cuoi not in content_normalized:
        return False, None, None
    
    lines = content.split('\n')
    best_result = None
    best_score = 0
    
    for i, line in enumerate(lines):
        line_stripped = line.strip()
        if not line_stripped:
            continue
            
        # Kiểm tra dòng có MSSV không
        if not mssv_pattern.search(line_stripped):
            continue
        
        line_normalized = normalize_text(line_stripped)
        parts = parse_table_row(line_stripped)
        
        stt = None
        nrl = None
        score = 1  # Base score vì có MSSV
        
        # === PHƯƠNG PHÁP 1: Dữ liệu trên cùng 1 dòng (bảng) ===
        if len(parts) >= 2:
            # Kiểm tra có tên trong dòng không
            if ten_normalized in line_normalized or ten_cuoi in line_normalized:
                score += 5  # Bonus lớn vì cùng dòng với tên
            
            # Tìm STT và NRL trong cùng dòng (dùng hàm mới)
            stt = find_stt_in_line(line_stripped, mssv)
            nrl = find_nrl_in_parts(parts, mssv)
            
            if stt:
                score += 2
            if nrl is not None:
                score += 3
        
        # === PHƯƠNG PHÁP 2: Dữ liệu trên nhiều dòng ===
        if stt is None or nrl is None:
            # Tìm trong vùng lân cận (5 dòng trước và sau)
            context_start = max(0, i - 5)
            context_end = min(len(lines), i + 6)
            context_lines = lines[context_start:context_end]
            context_text = normalize_text(' '.join(context_lines))
            
            # Kiểm tra tên có trong vùng lân cận không
            if ten_normalized in context_text or ten_cuoi in context_text:
                score += 2
            else:
                # Tên không gần MSSV -> có thể là người khác
                continue
            
            # Tìm STT ở các dòng trước hoặc trong chính dòng hiện tại
            if stt is None:
                # Thử tìm trong chính dòng hiện tại trước
                stt = find_stt_in_line(line_stripped, mssv)
                
                # Nếu không có, tìm ở các dòng trước
                if stt is None:
                    for offset in range(1, 6):
                        if i - offset >= 0:
                            prev_line = lines[i - offset].strip()
                            # Thử tìm STT trong dòng
                            found_stt = find_stt_in_line(prev_line, mssv)
                            if found_stt:
                                stt = found_stt
                                break
                            # Fallback: check nếu toàn dòng là số
                            if is_valid_stt(prev_line):
                                val = extract_stt_value(prev_line)
                                if val and val <= 10000:
                                    stt = val
                                    break
            
            # Tìm NRL ở các dòng sau
            if nrl is None:
                for offset in range(1, 5):
                    if i + offset < len(lines):
                        next_line = lines[i + offset].strip().replace(',', '.')
                        valid, val = is_valid_nrl(next_line)
                        if valid:
                            nrl = val
                            break
        
        # Cập nhật kết quả tốt nhất
        if score > best_score:
            best_score = score
            best_result = (stt, nrl)
    
    # Trả về kết quả tốt nhất
    if best_result:
        return True, best_result[0], best_result[1]
    
    # Fallback: tìm thấy MSSV nhưng không xác định được chi tiết
    # Kiểm tra lại tên có gần MSSV không
    for i, line in enumerate(lines):
        if mssv_pattern.search(line):
            context_start = max(0, i - 3)
            context_end = min(len(lines), i + 4)
            context_text = normalize_text(' '.join(lines[context_start:context_end]))
            if ten_normalized in context_text or ten_cuoi in context_text:
                return True, None, None
    
    return False, None, None


def process_doc(doc, ten_sv, mssv, session):
    link = doc["link"]
    doc_name = doc["name"]
    
    try:
        content = read_doc_text(link, session)
        if content is None:
            return None
        
        found, stt, nrl = find_student_in_content(content, ten_sv, mssv)
        
        if found:
            short_name = doc_name[:50] + "..." if len(doc_name) > 50 else doc_name
            return {
                "link": link,
                "doc_name": short_name or "File",
                "stt": stt if stt else "-",
                "nrl": nrl if nrl is not None else "-",
            }
        return None
    except Exception as e:
        print(f"[ERROR] {link}: {e}")
        return None


def create_excel(results, ten_sv, mssv, total_nrl):
    output_file = f"ket_qua_{mssv}.xlsx"
    
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Ket qua NRL"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws_out.merge_cells('A1:D1')
    ws_out['A1'] = "BAO CAO KET QUA DIEM REN LUYEN"
    ws_out['A1'].font = Font(bold=True, size=14)
    ws_out['A1'].alignment = center
    
    ws_out['A3'] = "Ho va ten:"
    ws_out['B3'] = ten_sv.upper()
    ws_out['A4'] = "MSSV:"
    ws_out['B4'] = mssv
    ws_out['A5'] = "Ngay xuat:"
    ws_out['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws_out['A7'] = "So file tim thay:"
    ws_out['B7'] = len(results)
    ws_out['A8'] = "TONG DIEM NRL:"
    ws_out['B8'] = total_nrl
    ws_out['A8'].font = Font(bold=True, color="0000FF")
    ws_out['B8'].font = Font(bold=True, color="0000FF")
    
    table_headers = ["#", "STT", "NRL", "Ten file", "Link"]
    for col, h in enumerate(table_headers, 1):
        cell = ws_out.cell(row=10, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    for idx, r in enumerate(results, 1):
        row_num = 10 + idx
        ws_out.cell(row=row_num, column=1, value=idx).alignment = center
        ws_out.cell(row=row_num, column=2, value=r["stt"]).alignment = center
        ws_out.cell(row=row_num, column=3, value=r["nrl"]).alignment = center
        ws_out.cell(row=row_num, column=4, value=r["doc_name"])
        ws_out.cell(row=row_num, column=5, value=r["link"])
        
        for col in range(1, 6):
            ws_out.cell(row=row_num, column=col).border = thin_border
    
    ws_out.column_dimensions['A'].width = 5
    ws_out.column_dimensions['B'].width = 8
    ws_out.column_dimensions['C'].width = 8
    ws_out.column_dimensions['D'].width = 50
    ws_out.column_dimensions['E'].width = 60
    
    wb_out.save(output_file)
    print(f"[INFO] Created: {output_file}")
    return output_file


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/health')
def health():
    docs = get_doc_links()
    return jsonify({
        "status": "ok",
        "excel_file": EXCEL_FILE,
        "excel_exists": os.path.exists(EXCEL_FILE),
        "total_docs": len(docs)
    })


@app.route('/search', methods=['POST'])
def search():
    try:
        ten_sv = request.form.get('ten_sv', '').strip()
        mssv = request.form.get('mssv', '').strip()
        
        if not ten_sv or not mssv:
            return jsonify({"error": "Vui long nhap ca ten VA MSSV"})
        
        unique_docs = get_doc_links()
        
        if not unique_docs:
            return jsonify({"error": "Khong tim thay file Excel hoac file rong"})
        
        results = []
        
        session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=MAX_WORKERS,
            pool_maxsize=MAX_WORKERS
        )
        session.mount('https://', adapter)
        session.headers.update({"User-Agent": "Mozilla/5.0"})
        
        print(f"[INFO] Scanning {len(unique_docs)} files for {ten_sv} - {mssv}")
        
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(process_doc, doc, ten_sv, mssv, session): doc 
                for doc in unique_docs
            }
            
            try:
                for future in as_completed(futures, timeout=25):
                    try:
                        result = future.result(timeout=1)
                        if result:
                            results.append(result)
                    except:
                        pass
            except:
                pass
        
        results.sort(key=lambda x: (x["stt"] if isinstance(x["stt"], int) else 9999))
        total_nrl = sum(r["nrl"] for r in results if isinstance(r["nrl"], (int, float)))
        
        print(f"[INFO] Found {len(results)} results, total NRL: {total_nrl}")
        
        return jsonify({
            "results": results,
            "total_nrl": total_nrl,
            "total_files": len(results),
            "ten_sv": ten_sv,
            "mssv": mssv
        })
    except Exception as e:
        print(f"[ERROR] Search failed: {e}")
        return jsonify({"error": f"Loi server: {str(e)}"}), 500


@app.route('/download', methods=['POST'])
def download():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Khong co du lieu"}), 400
            
        results = data.get('results', [])
        ten_sv = data.get('ten_sv', '')
        mssv = data.get('mssv', '')
        total_nrl = data.get('total_nrl', 0)
        
        if not results:
            return jsonify({"error": "Khong co ket qua de tai"}), 400
        
        excel_file = create_excel(results, ten_sv, mssv, total_nrl)
        return send_file(excel_file, as_attachment=True)
    except Exception as e:
        print(f"[ERROR] Download failed: {e}")
        return jsonify({"error": f"Loi tao file: {str(e)}"}), 500


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("DEBUG", "false").lower() == "true"
    app.run(host='0.0.0.0', port=port, debug=debug)
