import re
import requests
from openpyxl import load_workbook
from unidecode import unidecode

# Lấy 1 link docs để xem cấu trúc
EXCEL_FILE = "nrl.xlsx"
TEN_SV = "vo duc tien"
MSSV = "2254820130"

headers = {"User-Agent": "Mozilla/5.0"}

wb = load_workbook(EXCEL_FILE)
ws = wb.active

# Lấy vài link đầu tiên
doc_links = []
for row in ws.iter_rows():
    for cell in row:
        if cell.hyperlink and cell.hyperlink.target:
            link = cell.hyperlink.target
            if "docs.google.com/document" in link:
                doc_links.append(link)
                if len(doc_links) >= 5:
                    break
    if len(doc_links) >= 5:
        break

# Đọc và tìm dòng chứa tên/MSSV
def read_doc_text(url):
    doc_id = re.search(r"/d/([a-zA-Z0-9_-]+)", url).group(1)
    export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=txt"
    r = requests.get(export_url, headers=headers, timeout=15)
    if "accounts.google.com" in r.url or r.status_code != 200:
        return None
    return r.text

ten_clean = unidecode(TEN_SV.lower())

for link in doc_links:
    print(f"\n{'='*60}")
    print(f"Link: {link[:60]}...")
    
    content = read_doc_text(link)
    if not content:
        print("🔒 Không truy cập được")
        continue
    
    lines = content.split('\n')
    
    # Tìm dòng chứa tên/MSSV
    for i, line in enumerate(lines):
        line_clean = unidecode(line.lower())
        if ten_clean in line_clean or MSSV in line:
            print(f"\n✅ TÌM THẤY tại dòng {i+1}:")
            print(f">>> {line}")
            print(f"\nCác số trong dòng: {re.findall(r'\\d+', line)}")
            
            # In thêm vài dòng xung quanh để xem context
            print(f"\n--- Context (3 dòng trước/sau) ---")
            start = max(0, i-3)
            end = min(len(lines), i+4)
            for j in range(start, end):
                marker = ">>>" if j == i else "   "
                print(f"{marker} [{j+1}] {lines[j]}")
            break
    else:
        print("❌ Không tìm thấy tên/MSSV")
