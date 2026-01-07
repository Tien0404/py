import pandas as pd
from openpyxl import load_workbook

EXCEL_FILE = "nrl.xlsx"

# Cách 1: Đọc bằng pandas
print("=== PANDAS ===")
df = pd.read_excel(EXCEL_FILE)
print(f"Số dòng: {len(df)}, Số cột: {len(df.columns)}")
print(f"Tên cột: {list(df.columns)}")
print("\n5 dòng đầu:")
print(df.head())

# Cách 2: Kiểm tra hyperlink ẩn bằng openpyxl
print("\n=== HYPERLINKS ẨN ===")
wb = load_workbook(EXCEL_FILE)
ws = wb.active

hyperlinks = []
for row in ws.iter_rows():
    for cell in row:
        if cell.hyperlink:
            hyperlinks.append(cell.hyperlink.target)
            if len(hyperlinks) <= 10:
                print(f"Cell {cell.coordinate}: {cell.hyperlink.target}")

print(f"\nTổng số hyperlink tìm thấy: {len(hyperlinks)}")

# Kiểm tra có link docs không
docs_links = [h for h in hyperlinks if h and "docs.google.com" in h]
print(f"Số link Google Docs: {len(docs_links)}")
