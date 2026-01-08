import re
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from unidecode import unidecode
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading


# =====================
# TH√îNG TIN C·ª¶A B·∫†N
# =====================
TEN_SV = "Cao Ho√†ng Tr√≠"
MSSV = "2433520225"

EXCEL_FILE = "nrl.xlsx"
OUTPUT_XLSX = "ket_qua_nrl.xlsx"
MAX_WORKERS = 20
REQUEST_TIMEOUT = 8

http_headers = {"User-Agent": "Mozilla/5.0"}
session = requests.Session()
adapter = requests.adapters.HTTPAdapter(
    pool_connections=MAX_WORKERS,
    pool_maxsize=MAX_WORKERS
)
session.mount('https://', adapter)
session.headers.update(http_headers)

print_lock = threading.Lock()


# =====================
# ƒê·ªåC HYPERLINK T·ª™ EXCEL
# =====================
print("üìÇ ƒêang ƒë·ªçc file Excel...")
wb = load_workbook(EXCEL_FILE)
ws = wb.active

doc_links = []
for row in ws.iter_rows():
    for cell in row:
        if cell.hyperlink and cell.hyperlink.target:
            link = cell.hyperlink.target
            if "docs.google.com/document" in link:
                cell_value = str(cell.value) if cell.value else ""
                doc_links.append({"link": link, "name": cell_value})

seen = set()
unique_docs = []
for doc in doc_links:
    if doc["link"] not in seen:
        seen.add(doc["link"])
        unique_docs.append(doc)

print(f"üîé Ph√°t hi·ªán {len(unique_docs)} file Docs")
print(f"‚ö° ƒêang qu√©t v·ªõi {MAX_WORKERS} lu·ªìng song song...\n")


# =====================
# H√ÄM ƒê·ªåC N·ªòI DUNG DOCS
# =====================
def read_doc_text(url):
    try:
        match = re.search(r'/d/([a-zA-Z0-9_-]+)', url)
        if not match:
            return None
        doc_id = match.group(1)
        export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=txt"
        
        r = session.get(export_url, timeout=REQUEST_TIMEOUT)
        if r.status_code == 200 and "accounts.google.com" not in r.url:
            return r.text
        return None
    except:
        return None


def normalize_text(text):
    text = unidecode(text.lower())
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def is_valid_stt(s):
    return bool(re.match(r'^\d{1,4}$', s))


def is_valid_nrl(s):
    match = re.match(r'^(\d+\.?\d*)$', s)
    if match:
        val = float(match.group(1))
        return 0 <= val <= 10, val
    return False, None


def find_student_in_content(content, ten_sv, mssv):
    content_normalized = normalize_text(content)
    ten_normalized = normalize_text(ten_sv)
    
    has_name = ten_normalized in content_normalized
    has_mssv = mssv in content
    
    if not (has_name and has_mssv):
        return False, None, None
    
    lines = content.split('\n')
    
    for i, line in enumerate(lines):
        line_normalized = normalize_text(line)
        
        if ten_normalized in line_normalized:
            stt = None
            nrl = None
            
            if i > 0:
                prev = lines[i-1].strip()
                if is_valid_stt(prev):
                    stt = int(prev)
            
            for j in range(i+1, min(i+6, len(lines))):
                check = lines[j].strip().replace(',', '.')
                valid, val = is_valid_nrl(check)
                if valid:
                    nrl = val
                    break
            
            return True, stt, nrl
    
    for i, line in enumerate(lines):
        if mssv in line:
            stt = None
            nrl = None
            
            if i >= 3:
                stt_line = lines[i-3].strip()
                if is_valid_stt(stt_line):
                    stt = int(stt_line)
            
            if i + 1 < len(lines):
                nrl_line = lines[i+1].strip().replace(',', '.')
                valid, val = is_valid_nrl(nrl_line)
                if valid:
                    nrl = val
            
            return True, stt, nrl
    
    return True, None, None


def process_doc(doc, index, total, ten_sv, mssv):
    link = doc["link"]
    doc_name = doc["name"]
    
    try:
        content = read_doc_text(link)
        
        if content is None:
            with print_lock:
                print(f"[{index}/{total}] üîí Private/L·ªói")
            return None
        
        found, stt, nrl = find_student_in_content(content, ten_sv, mssv)
        
        if found:
            with print_lock:
                print(f"[{index}/{total}] ‚úÖ STT: {stt} | NRL: {nrl}")
            
            return {
                "link": link,
                "doc_name": doc_name,
                "stt": stt if stt else "N/A",
                "nrl": nrl if nrl is not None else "N/A",
            }
        else:
            with print_lock:
                print(f"[{index}/{total}] ‚ùå")
            return None

    except Exception as e:
        with print_lock:
            print(f"[{index}/{total}] ‚ö†Ô∏è L·ªói: {e}")
        return None


# =====================
# CH·∫†Y ƒêA LU·ªíNG
# =====================
results = []
total = len(unique_docs)

with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
    futures = {
        executor.submit(process_doc, doc, i, total, TEN_SV, MSSV): doc 
        for i, doc in enumerate(unique_docs, 1)
    }
    
    for future in as_completed(futures):
        try:
            result = future.result(timeout=1)
            if result:
                results.append(result)
        except:
            pass

total_nrl = sum(r["nrl"] for r in results if isinstance(r["nrl"], (int, float)))


# =====================
# XU·∫§T EXCEL
# =====================
def create_excel(results, ten_sv, mssv, total_nrl, output_file):
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Ket qua NRL"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws_out.merge_cells('A1:D1')
    ws_out['A1'] = "B√ÅO C√ÅO K·∫æT QU·∫¢ ƒêI·ªÇM R√àN LUY·ªÜN"
    ws_out['A1'].font = Font(bold=True, size=14)
    ws_out['A1'].alignment = center
    
    ws_out['A3'] = "H·ªç v√† t√™n:"
    ws_out['B3'] = ten_sv.upper()
    ws_out['A4'] = "MSSV:"
    ws_out['B4'] = mssv
    ws_out['A5'] = "Ng√†y xu·∫•t:"
    ws_out['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws_out['A7'] = "S·ªë file t√¨m th·∫•y:"
    ws_out['B7'] = len(results)
    ws_out['A8'] = "T·ªîNG ƒêI·ªÇM NRL:"
    ws_out['B8'] = total_nrl
    ws_out['A8'].font = Font(bold=True, color="0000FF")
    ws_out['B8'].font = Font(bold=True, color="0000FF")
    
    table_headers = ["#", "STT", "NRL", "T√™n file", "Link"]
    for col, h in enumerate(table_headers, 1):
        cell = ws_out.cell(row=10, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    for idx, r in enumerate(results, 1):
        row_num = 10 + idx
        doc_name = r["doc_name"] if r["doc_name"] else f"Link {idx}"
        
        ws_out.cell(row=row_num, column=1, value=idx).alignment = center
        ws_out.cell(row=row_num, column=2, value=r["stt"]).alignment = center
        ws_out.cell(row=row_num, column=3, value=r["nrl"]).alignment = center
        ws_out.cell(row=row_num, column=4, value=doc_name)
        ws_out.cell(row=row_num, column=5, value=r["link"])
        
        for col in range(1, 6):
            ws_out.cell(row=row_num, column=col).border = thin_border
    
    ws_out.column_dimensions['A'].width = 5
    ws_out.column_dimensions['B'].width = 10
    ws_out.column_dimensions['C'].width = 10
    ws_out.column_dimensions['D'].width = 40
    ws_out.column_dimensions['E'].width = 60
    
    try:
        wb_out.save(output_file)
        print(f"\nüìÑ ƒê√£ xu·∫•t Excel: {output_file}")
    except PermissionError:
        timestamp = datetime.now().strftime('%H%M%S')
        new_file = output_file.replace('.xlsx', f'_{timestamp}.xlsx')
        wb_out.save(new_file)
        print(f"\nüìÑ File c≈© ƒëang m·ªü, ƒë√£ xu·∫•t: {new_file}")


# =====================
# K·∫æT QU·∫¢
# =====================
print("\n" + "="*50)
print("üìã K·∫æT QU·∫¢ CU·ªêI C√ôNG")
print("="*50)

if results:
    print(f"‚úÖ T√¨m th·∫•y {len(results)} file ch·ª©a th√¥ng tin c·ªßa b·∫°n")
    print(f"üìä T·ªîNG ƒêI·ªÇM NRL: {total_nrl}\n")
    
    for idx, r in enumerate(results, 1):
        print(f"  {idx}. STT: {r['stt']} | NRL: {r['nrl']} | {r['doc_name'] or 'Link '+str(idx)}")
    
    create_excel(results, TEN_SV, MSSV, total_nrl, OUTPUT_XLSX)
else:
    print("‚ùå Kh√¥ng t√¨m th·∫•y file n√†o ch·ª©a t√™n/MSSV c·ªßa b·∫°n")
